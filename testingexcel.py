from openpyxl import load_workbook

filename = "BulkOperation_TEST.xlsx"
wb = load_workbook(filename=filename)

ws = wb['Template - Sponsored Product ']

#Varibales
keywords = []
asins = []

# Utility functions

def getAdGroup(campaign_name):
    x = campaign_name.split(sep="_")
    for item in x:
        if item == 'Market':
            ad_group_name = "_".join(x[x.index(item) + 1:])
            return ad_group_name

def getKeywords(filename="keywords.txt"):
    f = open(filename)
    f1 = f.readlines()
    for line in f1:
        line = line.strip('\n')
        keywords.append(line)

def getAsins(filename="asins.txt"):
    f = open(filename)
    f1 = f.readlines()
    for line in f1:
        line = line.strip('\n')
        asins.append(line)

# Actual Campaign Creation

lastRow = ws.max_row+1
def createCampaign(campaignName):
    global lastRow
    ws.cell(row=lastRow,column=2,value="Campaign")
    ws.cell(row=lastRow,column=4,value=campaignName)
    ws.cell(row=lastRow,column=5,value=15)
    ws.cell(row=lastRow,column=6,value="24/10/2019")
    ws.cell(row=lastRow,column=8,value="Manual")
    ws.cell(row=lastRow,column=16,value="Enabled")
    lastRow = lastRow + 1

def createAdGroup(campaignName):
    global lastRow
    ws.cell(row=lastRow,column=2,value="AdGroup")
    ws.cell(row=lastRow,column=4,value=campaignName)
    ws.cell(row=lastRow,column=10,value=getAdGroup(campaignName))
    ws.cell(row=lastRow,column=11,value=1)
    ws.cell(row=lastRow,column=17,value="Enabled")
    lastRow = lastRow + 1

def createAd(campaignName):
    getAsins()
    print(asins)
    global lastRow
    for asin in asins:
        ws.cell(row=lastRow,column=2,value="Ad")
        ws.cell(row=lastRow,column=4,value=campaignName)
        ws.cell(row=lastRow,column=10,value=getAdGroup(campaignName))
        ws.cell(row=lastRow,column=15,value=asin)
        ws.cell(row=lastRow,column=18,value="Enabled")
        lastRow = lastRow + 1

def createKeyword(campaignName):
    getKeywords()
    print(keywords)
    global lastRow
    for keyword in keywords:
        ws.cell(row=lastRow,column=2,value="Keyword")
        ws.cell(row=lastRow,column=4,value=campaignName)
        ws.cell(row=lastRow,column=10,value=getAdGroup(campaignName))
        ws.cell(row=lastRow,column=11,value=1)
        ws.cell(row=lastRow,column=12,value=keyword)
        ws.cell(row=lastRow,column=13,value="Broad")
        ws.cell(row=lastRow,column=18,value="Enabled")
        lastRow = lastRow + 1
    # creates brand keyword
    ws.cell(row=lastRow,column=2,value="Keyword")
    ws.cell(row=lastRow,column=4,value=campaignName)
    ws.cell(row=lastRow,column=10,value=getAdGroup(campaignName))
    ws.cell(row=lastRow,column=11,value=1)
    ws.cell(row=lastRow,column=12,value='ardes')
    ws.cell(row=lastRow,column=13,value="NegativePhrase")
    ws.cell(row=lastRow,column=18,value="Enabled")
    lastRow = lastRow + 1
    

def createTotalCampaign(campaign):
    print("creating campaign")
    createCampaign(campaign)
    print("creating ad group")
    createAdGroup(campaign)
    createAd(campaign)
    createKeyword(campaign)



createTotalCampaign("MEL_SP_Market_Adattatori_Viaggio")
#createAd("Ardes_SP_Market_Ha_Fottutamente_Funzionato")
#createKeyword("Ardes_SP_Market_Ha_Fottutamente_Funzionato")

wb.save(filename=filename)

# TODO - change datetime to timestamp, not string








